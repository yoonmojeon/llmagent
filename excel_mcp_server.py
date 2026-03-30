"""
Excel MCP Server — FastMCP 버전
Excel 파일 자동화를 위한 MCP 서버
"""

import os
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from fastmcp import FastMCP

mcp = FastMCP("excel-mcp-server")


@mcp.tool()
def read_excel(
    file_path: str,
    sheet_name: str = "",
    max_rows: int = 100,
) -> str:
    """Excel 파일을 읽어 내용을 반환합니다. 특정 시트를 지정할 수 있습니다.

    Args:
        file_path: 읽을 Excel 파일 경로 (절대 경로 또는 상대 경로)
        sheet_name: 읽을 시트 이름 (생략 시 첫 번째 시트)
        max_rows: 최대 읽을 행 수 (기본값: 100)
    """
    if not os.path.exists(file_path):
        return f"파일을 찾을 수 없습니다: {file_path}"
    sheet = sheet_name if sheet_name else 0
    df = pd.read_excel(file_path, sheet_name=sheet, nrows=max_rows)
    return f"파일: {file_path}\n크기: {df.shape[0]}행 × {df.shape[1]}열\n\n{df.to_string(index=True)}"


@mcp.tool()
def create_excel(
    file_path: str,
    headers: list[str] = [],
    data: list[list] = [],
    sheet_name: str = "Sheet1",
) -> str:
    """새 Excel 파일을 생성하고 데이터를 입력합니다.

    Args:
        file_path: 생성할 Excel 파일 경로
        headers: 헤더 행 (컬럼명 목록)
        data: 데이터 행들의 2D 배열 (예: [[값1, 값2], [값3, 값4]])
        sheet_name: 시트 이름 (기본값: Sheet1)
    """
    os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if headers:
        ws.append(headers)
    for row in data:
        ws.append(row)
    wb.save(file_path)
    row_count = len(data)
    col_count = len(headers) if headers else (len(data[0]) if data else 0)
    return f"Excel 파일 생성 완료: {file_path}\n시트: {sheet_name}\n데이터: {row_count}행 × {col_count}열"


@mcp.tool()
def update_cell(
    file_path: str,
    cell: str,
    value: str,
    sheet_name: str = "",
) -> str:
    """Excel 파일의 특정 셀 값을 업데이트합니다.

    Args:
        file_path: Excel 파일 경로
        cell: 셀 주소 (예: A1, B3, C10)
        value: 입력할 값
        sheet_name: 시트 이름 (생략 시 첫 번째 시트)
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    ws[cell] = value
    wb.save(file_path)
    return f"셀 {cell} → '{value}' 업데이트 완료 (파일: {file_path})"


@mcp.tool()
def get_cell_value(
    file_path: str,
    cell: str,
    sheet_name: str = "",
) -> str:
    """Excel 파일에서 특정 셀의 값을 읽어옵니다.

    Args:
        file_path: Excel 파일 경로
        cell: 읽을 셀 주소 (예: A1, B3)
        sheet_name: 시트 이름 (생략 시 첫 번째 시트)
    """
    if not os.path.exists(file_path):
        return f"파일을 찾을 수 없습니다: {file_path}"
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    return f"셀 {cell} 값: {ws[cell].value}"


@mcp.tool()
def get_sheet_names(file_path: str) -> str:
    """Excel 파일에 있는 모든 시트 이름을 반환합니다.

    Args:
        file_path: Excel 파일 경로
    """
    if not os.path.exists(file_path):
        return f"파일을 찾을 수 없습니다: {file_path}"
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames
    return f"시트 목록 ({len(sheets)}개): {', '.join(sheets)}"


@mcp.tool()
def add_sheet(file_path: str, sheet_name: str) -> str:
    """Excel 파일에 새 시트를 추가합니다.

    Args:
        file_path: Excel 파일 경로
        sheet_name: 추가할 시트 이름
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
    else:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            return f"시트 '{sheet_name}'이 이미 존재합니다."
        wb.create_sheet(sheet_name)
    wb.save(file_path)
    return f"시트 '{sheet_name}' 추가 완료 (파일: {file_path})"


@mcp.tool()
def rename_sheet(file_path: str, old_name: str, new_name: str) -> str:
    """Excel 파일의 시트 이름을 변경합니다.

    Args:
        file_path: Excel 파일 경로
        old_name: 변경할 시트의 현재 이름
        new_name: 새 시트 이름
    """
    if not os.path.exists(file_path):
        return f"파일을 찾을 수 없습니다: {file_path}"
    wb = openpyxl.load_workbook(file_path)
    if old_name not in wb.sheetnames:
        return f"시트 '{old_name}'을 찾을 수 없습니다."
    wb[old_name].title = new_name
    wb.save(file_path)
    return f"시트 이름 변경: '{old_name}' → '{new_name}'"


@mcp.tool()
def write_data_to_sheet(
    file_path: str,
    sheet_name: str,
    data: list[list],
    start_row: int = 1,
    start_col: int = 1,
) -> str:
    """Excel 파일의 특정 시트에 데이터를 일괄 입력합니다.

    Args:
        file_path: Excel 파일 경로
        sheet_name: 데이터를 입력할 시트 이름
        data: 입력할 데이터 2D 배열 (예: [[값1, 값2], [값3, 값4]])
        start_row: 시작 행 번호 (1부터, 기본값: 1)
        start_col: 시작 열 번호 (1부터, 기본값: 1)
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.active.title = sheet_name
    else:
        wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
    wb.save(file_path)
    return f"{len(data)}행 데이터를 '{sheet_name}' 시트에 입력 완료 (시작: 행{start_row}, 열{start_col})"


@mcp.tool()
def delete_rows(
    file_path: str,
    row_start: int,
    row_end: int = 0,
    sheet_name: str = "",
) -> str:
    """Excel 파일에서 특정 행(들)을 삭제합니다.

    Args:
        file_path: Excel 파일 경로
        row_start: 삭제 시작 행 번호
        row_end: 삭제 종료 행 번호 (0이면 row_start만 삭제)
        sheet_name: 시트 이름 (생략 시 첫 번째 시트)
    """
    if not os.path.exists(file_path):
        return f"파일을 찾을 수 없습니다: {file_path}"
    end = row_end if row_end else row_start
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    count = end - row_start + 1
    ws.delete_rows(row_start, count)
    wb.save(file_path)
    return f"행 {row_start}~{end} ({count}행) 삭제 완료"


@mcp.tool()
def apply_formula(
    file_path: str,
    cell: str,
    formula: str,
    sheet_name: str = "",
) -> str:
    """Excel 셀에 수식을 적용합니다.

    Args:
        file_path: Excel 파일 경로
        cell: 수식을 입력할 셀 주소 (예: A10)
        formula: Excel 수식 (예: =SUM(A1:A9), =AVERAGE(B2:B8))
        sheet_name: 시트 이름 (생략 시 첫 번째 시트)
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    ws[cell] = formula
    wb.save(file_path)
    return f"수식 '{formula}' → 셀 {cell} 적용 완료"


if __name__ == "__main__":
    mcp.run(show_banner=False)
